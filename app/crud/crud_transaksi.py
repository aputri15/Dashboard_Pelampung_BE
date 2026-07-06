from typing import Optional, List, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, inspect, text
from sqlalchemy.exc import IntegrityError
from app.models.transaksi import Transaksi, LogUpload
from app.schemas.transaksi import TransaksiCreate, TransaksiUpdate
from datetime import datetime
import hashlib
import io
import re

# ==================== TRANSAKSI CRUD ====================

def get_transaksi(db: Session, transaksi_id: int) -> Optional[Transaksi]:
    return db.query(Transaksi).filter(Transaksi.id == transaksi_id).first()

def get_transaksi_list(
    db: Session,
    page: int = 1,
    per_page: int = 15,
    search: str = None,
    wilayah: str = None,
    bulan: str = None,
    tahun: str = None,
) -> Tuple[List[Transaksi], int]:
    """Get paginated and filtered transaksi list."""
    query = db.query(Transaksi)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Transaksi.nama_pelanggan.ilike(search_term),
                Transaksi.kategori.ilike(search_term),
                Transaksi.nama_model.ilike(search_term),
                Transaksi.kota.ilike(search_term),
            )
        )

    if wilayah and wilayah != "Semua":
        query = query.filter(Transaksi.wilayah == wilayah)

    if tahun and tahun != "Semua":
        query = query.filter(Transaksi.tanggal_po.like(f"{tahun}-%"))

    if bulan and bulan != "Semua":
        bulan_value = str(bulan).zfill(2)
        query = query.filter(func.substr(Transaksi.tanggal_po, 6, 2) == bulan_value)

    total = query.count()
    skip = (page - 1) * per_page
    data = query.order_by(Transaksi.id.desc()).offset(skip).limit(per_page).all()

    return data, total

def get_all_wilayah(db: Session) -> List[str]:
    """Get distinct wilayah values for filter dropdown."""
    results = db.query(Transaksi.wilayah).distinct().all()
    return [r[0] for r in results if r[0]]

def create_transaksi(db: Session, transaksi_in: TransaksiCreate) -> Transaksi:
    db_transaksi = Transaksi(**transaksi_in.model_dump())
    db.add(db_transaksi)
    db.commit()
    db.refresh(db_transaksi)
    return db_transaksi

def create_manual_transaksi(db: Session, transaksi_in) -> Transaksi:
    data = transaksi_in.model_dump()
    data["total_harga"] = float(data["qty"]) * float(data["harga_satuan"])
    db_transaksi = Transaksi(**data)
    db.add(db_transaksi)
    db.commit()
    db.refresh(db_transaksi)
    return db_transaksi

def update_transaksi(db: Session, db_transaksi: Transaksi, transaksi_in: TransaksiUpdate) -> Transaksi:
    update_data = transaksi_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_transaksi, field, value)
    if "qty" in update_data or "harga_satuan" in update_data:
        qty = db_transaksi.qty or 0
        harga_satuan = db_transaksi.harga_satuan or 0
        db_transaksi.total_harga = float(qty) * float(harga_satuan)
    db.add(db_transaksi)
    db.commit()
    db.refresh(db_transaksi)
    return db_transaksi

def get_transaksi_filter_options(db: Session) -> dict:
    dates = db.query(Transaksi.tanggal_po).distinct().all()
    bulan_set = set()
    tahun_set = set()
    for (tanggal,) in dates:
        if tanggal and len(tanggal) >= 7 and "-" in tanggal:
            tahun_set.add(tanggal[:4])
            bulan_set.add(tanggal[5:7])

    wilayah = db.query(Transaksi.wilayah).distinct().all()
    wilayah_list = sorted([row[0] for row in wilayah if row[0]])

    return {
        "bulan": sorted(bulan_set),
        "tahun": sorted(tahun_set),
        "wilayah": wilayah_list,
    }

def delete_transaksi(db: Session, transaksi_id: int) -> Transaksi:
    db_transaksi = db.query(Transaksi).filter(Transaksi.id == transaksi_id).first()
    if db_transaksi:
        db.delete(db_transaksi)
        db.commit()
    return db_transaksi

def delete_all_transaksi(db: Session) -> int:
    """Delete all transaksi records. Returns count of deleted rows."""
    count = db.query(Transaksi).count()
    db.query(Transaksi).delete()
    db.commit()
    return count

def get_transaksi_stats(db: Session) -> dict:
    """Get summary statistics for transaksi data."""
    total_rows = db.query(Transaksi).count()
    total_revenue = db.query(func.sum(Transaksi.total_harga)).scalar() or 0
    total_qty = db.query(func.sum(Transaksi.qty)).scalar() or 0
    unique_customers = db.query(func.count(func.distinct(Transaksi.nama_pelanggan))).scalar() or 0
    return {
        "total_rows": total_rows,
        "total_revenue": total_revenue,
        "total_qty": total_qty,
        "unique_customers": unique_customers,
    }


SUCCESS_UPLOAD_STATUSES = ("Sukses", "Berhasil")


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _log_can_allow_reupload(log: LogUpload) -> bool:
    return (
        log is not None
        and log.file_hash is not None
        and log.status in SUCCESS_UPLOAD_STATUSES
        and not bool(log.reupload_allowed)
        and not bool(log.reupload_used_at)
    )


def _find_reupload_permission_target(db: Session, log: LogUpload) -> Optional[LogUpload]:
    if _log_can_allow_reupload(log):
        return log
    if log is None or not log.file_hash:
        return None
    return (
        db.query(LogUpload)
        .filter(LogUpload.file_hash == log.file_hash)
        .filter(LogUpload.status.in_(SUCCESS_UPLOAD_STATUSES))
        .filter(LogUpload.reupload_allowed == False)  # noqa: E712
        .filter(LogUpload.reupload_used_at.is_(None))
        .order_by(LogUpload.id.desc())
        .first()
    )


def _decorate_log_upload(log: LogUpload, db: Optional[Session] = None) -> LogUpload:
    can_allow_reupload = _log_can_allow_reupload(log)
    if not can_allow_reupload and db is not None:
        can_allow_reupload = _find_reupload_permission_target(db, log) is not None
    setattr(log, "can_allow_reupload", can_allow_reupload)
    return log


def _format_error_value(value) -> str:
    if value is None:
        return ""
    return str(value)


def _upload_error(row: int, field: str, value, reason: str, fix: str) -> dict:
    return {
        "row": row,
        "column": field,
        "field": field,
        "value": _format_error_value(value),
        "reason": reason,
        "fix": fix,
    }


def calculate_file_hash(file_content: bytes) -> str:
    return hashlib.sha256(file_content).hexdigest()


def ensure_log_upload_columns(db: Session) -> None:
    bind = db.get_bind()
    inspector = inspect(bind)
    has_transaksi = inspector.has_table("transaksi")
    has_log_upload = inspector.has_table("log_upload")
    transaksi_columns = (
        {column["name"] for column in inspector.get_columns("transaksi")}
        if has_transaksi
        else set()
    )
    log_upload_columns = (
        {column["name"] for column in inspector.get_columns("log_upload")}
        if has_log_upload
        else set()
    )
    with bind.begin() as connection:
        if has_transaksi:
            if "source_log_id" not in transaksi_columns:
                connection.execute(text("ALTER TABLE transaksi ADD COLUMN source_log_id INTEGER"))
            if "uploaded_at" not in transaksi_columns:
                connection.execute(text("ALTER TABLE transaksi ADD COLUMN uploaded_at VARCHAR"))
            connection.execute(
                text("CREATE INDEX IF NOT EXISTS ix_transaksi_source_log_id ON transaksi (source_log_id)")
            )

        if not has_log_upload:
            return

        if "file_hash" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN file_hash VARCHAR"))
        if "is_deleted" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN is_deleted BOOLEAN NOT NULL DEFAULT 0"))
        if "reupload_allowed" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN reupload_allowed BOOLEAN NOT NULL DEFAULT 0"))
        if "reupload_allowed_by" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN reupload_allowed_by VARCHAR"))
        if "reupload_allowed_at" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN reupload_allowed_at VARCHAR"))
        if "reupload_used_at" not in log_upload_columns:
            connection.execute(text("ALTER TABLE log_upload ADD COLUMN reupload_used_at VARCHAR"))
        connection.execute(
            text("CREATE INDEX IF NOT EXISTS ix_log_upload_file_hash ON log_upload (file_hash)")
        )
        connection.execute(text("DROP INDEX IF EXISTS ux_log_upload_success_file_hash"))
        connection.execute(
            text(
                "CREATE UNIQUE INDEX IF NOT EXISTS ux_log_upload_active_success_file_hash "
                "ON log_upload (file_hash) "
                "WHERE file_hash IS NOT NULL "
                "AND status IN ('Sukses', 'Berhasil') "
                "AND reupload_allowed = 0 "
                "AND reupload_used_at IS NULL"
            )
        )


def get_successful_upload_by_hash(db: Session, file_hash: str) -> Optional[LogUpload]:
    ensure_log_upload_columns(db)
    return _query_successful_upload_by_hash(db, file_hash)


def _query_successful_upload_by_hash(db: Session, file_hash: str) -> Optional[LogUpload]:
    return (
        db.query(LogUpload)
        .filter(LogUpload.file_hash == file_hash)
        .filter(LogUpload.status.in_(SUCCESS_UPLOAD_STATUSES))
        .order_by(LogUpload.id.desc())
        .first()
    )


def get_unused_reupload_permission_by_hash(db: Session, file_hash: str) -> Optional[LogUpload]:
    ensure_log_upload_columns(db)
    return (
        db.query(LogUpload)
        .filter(LogUpload.file_hash == file_hash)
        .filter(LogUpload.status.in_(SUCCESS_UPLOAD_STATUSES))
        .filter(LogUpload.reupload_allowed == True)  # noqa: E712
        .filter(LogUpload.reupload_used_at.is_(None))
        .order_by(LogUpload.id.desc())
        .first()
    )


def build_duplicate_upload_response(duplicate: LogUpload) -> dict:
    message = (
        f"File ini pernah di-upload pada {duplicate.tanggal} oleh {duplicate.uploaded_by}. "
        "Jika ingin upload ulang, buka Riwayat Log Upload lalu klik Izinkan Upload Ulang."
    )
    return {
        "success": False,
        "total_rows": 0,
        "inserted_rows": 0,
        "skipped_rows": 0,
        "error_code": "duplicate_upload",
        "duplicate_upload": {
            "id": duplicate.id,
            "tanggal": duplicate.tanggal,
            "nama_file": duplicate.nama_file,
            "uploaded_by": duplicate.uploaded_by,
        },
        "errors": [message],
        "error_details": [],
        "message": message,
    }


# ==================== LOG UPLOAD CRUD ====================

def get_log_uploads(
    db: Session,
    page: int = 1,
    per_page: int = 50,
    search: str = None,
    bulan: str = None,
    tahun: str = None,
    status: str = None,
    include_deleted: bool = False,
) -> Tuple[List[LogUpload], int]:
    ensure_log_upload_columns(db)
    query = db.query(LogUpload)

    if not include_deleted:
        query = query.filter(LogUpload.is_deleted == False)  # noqa: E712

    if search:
        query = query.filter(LogUpload.nama_file.ilike(f"%{search}%"))

    if tahun and tahun != "Semua":
        query = query.filter(LogUpload.tanggal.like(f"{tahun}-%"))

    if bulan and bulan != "Semua":
        bulan_value = str(bulan).zfill(2)
        query = query.filter(func.substr(LogUpload.tanggal, 6, 2) == bulan_value)

    if status and status != "Semua":
        query = query.filter(LogUpload.status == status)

    total = query.count()
    skip = (page - 1) * per_page
    logs = query.order_by(LogUpload.id.desc()).offset(skip).limit(per_page).all()
    return [_decorate_log_upload(log, db) for log in logs], total

def get_log_upload_count(db: Session) -> int:
    ensure_log_upload_columns(db)
    return db.query(LogUpload).filter(LogUpload.is_deleted == False).count()  # noqa: E712

def get_log_upload_filter_options(db: Session) -> dict:
    ensure_log_upload_columns(db)
    rows = db.query(LogUpload.tanggal, LogUpload.status).filter(LogUpload.is_deleted == False).all()  # noqa: E712
    bulan_set = set()
    tahun_set = set()
    status_set = set()
    for tanggal, status in rows:
        if tanggal and len(tanggal) >= 7 and "-" in tanggal:
            tahun_set.add(tanggal[:4])
            bulan_set.add(tanggal[5:7])
        if status:
            status_set.add(status)
    return {
        "bulan": sorted(bulan_set),
        "tahun": sorted(tahun_set),
        "status": sorted(status_set),
    }

def create_log_upload(
    db: Session,
    nama_file: str,
    jumlah_baris: int,
    status: str,
    uploaded_by: str,
    file_hash: Optional[str] = None,
    commit: bool = True,
) -> LogUpload:
    ensure_log_upload_columns(db)
    log = LogUpload(
        tanggal=_now_str(),
        nama_file=nama_file,
        jumlah_baris=jumlah_baris,
        status=status,
        uploaded_by=uploaded_by,
        file_hash=file_hash,
    )
    db.add(log)
    if commit:
        db.commit()
        db.refresh(log)
    return _decorate_log_upload(log, db)

def delete_log_upload(db: Session, log_id: int) -> Optional[LogUpload]:
    ensure_log_upload_columns(db)
    log = db.query(LogUpload).filter(LogUpload.id == log_id).first()
    if log:
        db.delete(log)
        db.commit()
    return _decorate_log_upload(log, db) if log else None

def soft_delete_log_upload(db: Session, log_id: int) -> Optional[LogUpload]:
    ensure_log_upload_columns(db)
    log = db.query(LogUpload).filter(LogUpload.id == log_id).first()
    if log:
        log.is_deleted = True
        db.add(log)
        db.commit()
        db.refresh(log)
    return _decorate_log_upload(log, db) if log else None


def allow_log_reupload(db: Session, log_id: int, allowed_by: str) -> Optional[LogUpload]:
    ensure_log_upload_columns(db)
    log = db.query(LogUpload).filter(LogUpload.id == log_id).first()
    target_log = _find_reupload_permission_target(db, log)
    if not target_log:
        return None

    target_log.reupload_allowed = True
    target_log.reupload_allowed_by = allowed_by
    target_log.reupload_allowed_at = _now_str()
    db.add(target_log)
    db.commit()
    db.refresh(target_log)
    return _decorate_log_upload(target_log, db)


# ==================== EXCEL UPLOAD PROCESSING ====================

# Column name mapping: expected DB field -> possible Excel header variations
COLUMN_MAP = {
    "nomor_po": ["nomor po", "no po", "nomor_po", "no_po", "po number", "po"],
    "tanggal_po": ["tanggal po", "tanggal_po", "tanggal", "tgl po", "tgl_po", "date"],
    "id_pelanggan": ["id pelanggan", "id_pelanggan", "customer id", "id customer"],
    "nama_pelanggan": ["nama pelanggan", "nama_pelanggan", "pelanggan", "customer", "customer name"],
    "wilayah": ["wilayah", "region", "area"],
    "provinsi": ["provinsi", "province"],
    "kota": ["kota", "city", "kabupaten"],
    "id_produk": ["id produk", "id_produk", "product id", "kode produk"],
    "nama_model": ["nama model", "nama_model", "model", "produk", "product", "nama produk"],
    "kategori": ["kategori", "category", "jenis"],
    "qty": ["qty", "quantity", "jumlah", "kuantitas"],
    "harga_satuan": ["harga satuan", "harga_satuan", "unit price", "harga", "price"],
    "total_harga": ["total harga", "total_harga", "total", "subtotal", "total price"],
    "modal_unit": ["modal unit", "modal_unit", "modal", "cost", "harga modal"],
}

MASTER_SHEET_RE = re.compile(r"^MASTER\d{4}$")

REQUIRED_HEADERS = [
    "nomor_po",
    "tanggal_po",
    "id_pelanggan",
    "nama_pelanggan",
    "wilayah",
    "provinsi",
    "kota",
    "id_produk",
    "nama_model",
    "kategori",
    "qty",
    "harga_satuan",
    "total_harga",
    "modal_unit",
]

COMPUTED_FIELDS = {"total_harga"}
REQUIRED_VALUE_FIELDS = [field for field in REQUIRED_HEADERS if field not in COMPUTED_FIELDS]
FORWARD_FILL_FIELDS = [
    "nomor_po",
    "tanggal_po",
    "nama_pelanggan",
    "id_pelanggan",
    "wilayah",
    "provinsi",
    "kota",
]
TEXT_FIELDS = [
    "nomor_po",
    "tanggal_po",
    "id_pelanggan",
    "nama_pelanggan",
    "wilayah",
    "provinsi",
    "kota",
    "id_produk",
    "nama_model",
    "kategori",
]
NUMERIC_FIELDS = ["harga_satuan", "total_harga", "modal_unit"]

def _match_column(header: str) -> Optional[str]:
    """Match an Excel header to a DB field name using the COLUMN_MAP."""
    header_lower = header.strip().lower()
    for field, aliases in COLUMN_MAP.items():
        if header_lower in aliases:
            return field
    return None


def _is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _parse_float(value):
    if _is_empty(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text_value = str(value).strip().replace(" ", "")
    if "," in text_value and "." in text_value:
        if text_value.rfind(",") > text_value.rfind("."):
            text_value = text_value.replace(".", "").replace(",", ".")
        else:
            text_value = text_value.replace(",", "")
    elif "," in text_value:
        parts = text_value.split(",")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
            text_value = "".join(parts)
        else:
            text_value = text_value.replace(",", ".")
    elif "." in text_value:
        parts = text_value.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
            text_value = "".join(parts)
    return float(text_value)


def _parse_int(value):
    parsed = _parse_float(value)
    if parsed is None:
        return None
    return int(parsed)


def _select_master_sheet(wb):
    master_sheets = [sheet_name for sheet_name in wb.sheetnames if MASTER_SHEET_RE.fullmatch(sheet_name)]
    if len(master_sheets) == 0:
        return None, {
            "success": False,
            "total_rows": 0,
            "inserted_rows": 0,
            "skipped_rows": 0,
            "errors": ["Sheet transaksi tidak ditemukan. Nama sheet harus MASTER + 4 digit tahun, contoh MASTER2025."],
            "error_details": [],
            "message": "Sheet transaksi harus bernama MASTER + 4 digit tahun, contoh MASTER2025.",
        }
    if len(master_sheets) > 1:
        return None, {
            "success": False,
            "total_rows": 0,
            "inserted_rows": 0,
            "skipped_rows": 0,
            "errors": [f"File memiliki lebih dari satu sheet transaksi: {', '.join(master_sheets)}."],
            "error_details": [],
            "message": "File memiliki lebih dari satu sheet transaksi MASTERyyyy.",
        }
    return wb[master_sheets[0]], None


def process_excel_upload(db: Session, file_content: bytes, filename: str, uploaded_by: str) -> dict:
    """
    Parse an uploaded Excel file and insert rows into the transaksi table.
    Returns a dict with status info: { success, total_rows, inserted_rows, errors, message }
    """
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "total_rows": 0,
            "inserted_rows": 0,
            "errors": ["openpyxl not installed"],
            "error_details": [],
            "message": "Server error: openpyxl library not available.",
        }

    file_hash = calculate_file_hash(file_content)

    try:
        ensure_log_upload_columns(db)

        reupload_permission = get_unused_reupload_permission_by_hash(db, file_hash)
        duplicate = None if reupload_permission else get_successful_upload_by_hash(db, file_hash)
        if duplicate:
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return build_duplicate_upload_response(duplicate)

        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws, sheet_error = _select_master_sheet(wb)
        if sheet_error:
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return sheet_error

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return {
                "success": False,
                "total_rows": 0,
                "inserted_rows": 0,
                "skipped_rows": 0,
                "errors": ["File kosong atau hanya memiliki header."],
                "error_details": [],
                "message": "File tidak mengandung data.",
            }

        raw_headers = [str(h).strip() if h else "" for h in rows[0]]
        mapped_headers = {}
        for idx, h in enumerate(raw_headers):
            field = _match_column(h)
            if field:
                mapped_headers[idx] = field

        mapped_fields = set(mapped_headers.values())
        missing_headers = [field for field in REQUIRED_HEADERS if field not in mapped_fields]
        if missing_headers:
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return {
                "success": False,
                "total_rows": len(rows) - 1,
                "inserted_rows": 0,
                "skipped_rows": len(rows) - 1,
                "columns_detected": sorted(mapped_fields),
                "columns_not_mapped": [h for h in raw_headers if h and not _match_column(h)],
                "errors": [f"Kolom wajib tidak ditemukan: {', '.join(missing_headers)}"],
                "error_details": [],
                "message": "Kolom wajib tidak ditemukan dalam file.",
            }

        data_rows = rows[1:]
        errors = []
        error_details = []
        valid_rows = []
        ff_last = {}
        blank_row_count = 0
        blank_rows = []
        total_data_rows = 0

        for row_idx, row in enumerate(data_rows, start=2):
            row_data = {}
            for col_idx, field in mapped_headers.items():
                row_data[field] = row[col_idx] if col_idx < len(row) else None

            if all(_is_empty(value) for value in row_data.values()):
                blank_row_count += 1
                blank_rows.append(row_idx)
                continue
            total_data_rows += 1

            for ff_field in FORWARD_FILL_FIELDS:
                if _is_empty(row_data.get(ff_field)):
                    if ff_field in ff_last:
                        row_data[ff_field] = ff_last[ff_field]
                else:
                    ff_last[ff_field] = row_data[ff_field]

            row_errors = []
            row_error_details = []

            try:
                row_data["qty"] = _parse_int(row_data.get("qty"))
            except (ValueError, TypeError):
                raw_value = row_data.get("qty")
                row_errors.append(f"qty bukan angka valid: {raw_value}")
                row_error_details.append(
                    _upload_error(
                        row_idx,
                        "qty",
                        raw_value,
                        "Nilai qty bukan angka valid.",
                        "isi qty dengan angka bulat lebih dari 0.",
                    )
                )
                row_data["qty"] = None

            for field in NUMERIC_FIELDS:
                try:
                    row_data[field] = _parse_float(row_data.get(field))
                except (ValueError, TypeError):
                    raw_value = row_data.get(field)
                    row_errors.append(f"{field} bukan angka valid: {raw_value}")
                    row_error_details.append(
                        _upload_error(
                            row_idx,
                            field,
                            raw_value,
                            f"Nilai {field} bukan angka valid.",
                            f"isi {field} dengan angka tanpa huruf atau simbol selain pemisah ribuan/desimal.",
                        )
                    )
                    row_data[field] = None

            for field in TEXT_FIELDS:
                if not _is_empty(row_data.get(field)):
                    row_data[field] = str(row_data[field]).strip()

            empty_fields = [field for field in REQUIRED_VALUE_FIELDS if _is_empty(row_data.get(field))]
            if empty_fields:
                row_errors.append(f"kolom kosong: {', '.join(empty_fields)}")
                for field in empty_fields:
                    row_error_details.append(
                        _upload_error(
                            row_idx,
                            field,
                            row_data.get(field),
                            f"Kolom {field} kosong.",
                            f"isi {field} sesuai data transaksi pada spreadsheet.",
                        )
                    )

            qty_val = row_data.get("qty")
            harga_val = row_data.get("harga_satuan")
            total_val = row_data.get("total_harga")
            if qty_val is not None and harga_val is not None:
                expected_total = float(qty_val) * float(harga_val)
                if total_val is None:
                    row_data["total_harga"] = expected_total
                elif abs(float(total_val) - expected_total) > 1:
                    row_errors.append(
                        f"total_harga tidak sesuai rumus qty * harga_satuan: {total_val} != {expected_total}"
                    )
                    row_error_details.append(
                        _upload_error(
                            row_idx,
                            "total_harga",
                            total_val,
                            f"Total harga tidak sesuai rumus qty * harga_satuan ({expected_total}).",
                            "Samakan total_harga dengan qty * harga_satuan atau kosongkan agar dihitung otomatis.",
                        )
                    )

            if row_errors:
                errors.append(f"Baris {row_idx}: {'; '.join(row_errors)}.")
                error_details.extend(row_error_details)
                continue

            valid_rows.append({key: value for key, value in row_data.items() if value is not None})

        if errors:
            db.rollback()
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return {
                "success": False,
                "total_rows": total_data_rows,
                "inserted_rows": 0,
                "skipped_rows": total_data_rows,
                "columns_detected": sorted(mapped_fields),
                "columns_not_mapped": [h for h in raw_headers if h and not _match_column(h)],
                "errors": errors[:20],
                "error_details": error_details[:50],
                "message": "Upload gagal. Lengkapi atau perbaiki data Excel terlebih dahulu.",
            }

        if not valid_rows:
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return {
                "success": False,
                "total_rows": total_data_rows,
                "inserted_rows": 0,
                "skipped_rows": total_data_rows,
                "columns_detected": sorted(mapped_fields),
                "columns_not_mapped": [h for h in raw_headers if h and not _match_column(h)],
                "errors": ["File tidak mengandung baris transaksi valid."],
                "error_details": [],
                "message": "File tidak mengandung data transaksi valid.",
            }

        inserted = len(valid_rows)
        try:
            upload_timestamp = _now_str()
            success_log = LogUpload(
                tanggal=upload_timestamp,
                nama_file=filename,
                jumlah_baris=inserted,
                status="Sukses",
                uploaded_by=uploaded_by,
                file_hash=file_hash,
                is_deleted=False,
                reupload_allowed=False,
            )
            db.add(success_log)
            db.flush()

            for row_data in valid_rows:
                row_data["source_log_id"] = success_log.id
                row_data["uploaded_at"] = upload_timestamp
                db.add(Transaksi(**row_data))

            if reupload_permission:
                reupload_permission.reupload_used_at = upload_timestamp
                db.add(reupload_permission)

            db.commit()
            db.refresh(success_log)
        except IntegrityError:
            db.rollback()
            duplicate = _query_successful_upload_by_hash(db, file_hash)
            if duplicate:
                create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
                return build_duplicate_upload_response(duplicate)
            create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
            return {
                "success": False,
                "total_rows": 0,
                "inserted_rows": 0,
                "skipped_rows": 0,
                "error_code": "upload_conflict",
                "errors": ["Upload gagal karena konflik penyimpanan. Coba ulang beberapa saat lagi."],
                "error_details": [],
                "message": "Upload gagal karena konflik penyimpanan.",
            }

        message = f"Berhasil mengimpor {inserted} dari {total_data_rows} baris."
        if blank_row_count:
            message += f" Baris kosong dilewati: {', '.join(str(row) for row in blank_rows)}."

        return {
            "success": True,
            "total_rows": total_data_rows,
            "processed_rows": total_data_rows,
            "inserted_rows": inserted,
            "skipped_rows": blank_row_count,
            "blank_row_count": blank_row_count,
            "blank_rows": blank_rows,
            "columns_detected": sorted(mapped_fields),
            "columns_not_mapped": [h for h in raw_headers if h and not _match_column(h)],
            "errors": [],
            "error_details": [],
            "message": message,
        }

    except Exception as e:
        db.rollback()
        create_log_upload(db, filename, 0, "Gagal", uploaded_by, file_hash=file_hash)
        return {
            "success": False,
            "total_rows": 0,
            "inserted_rows": 0,
            "errors": [str(e)],
            "error_details": [],
            "message": f"Error saat memproses file: {str(e)}",
        }
